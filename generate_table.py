"""Library to convert E-Redes readings into a Pandas Dataframe"""

import openpyxl
import pandas as pd #pip install Pyarrow
import os


def open_reading(path):
    """Reads a .xlsx given by E-Redes and organizes the values into a list.

    Args:
        path (String): Path to the .xlsx file

    Returns:
        List<List<String>>: Values of the Excel file
    """
    sheet = openpyxl.load_workbook(path).active
    rows = sheet.max_row
    data = []
    
    for line in range(9, rows, 2):
        date = sheet.cell(line, 1).value
        ccv = (sheet.cell(line, 5).value)       # energia Consumida no Contador em Vazio
        csv = (sheet.cell(line, 6).value)       # energia Consumida no Saldo em Vazio
        ccp = (sheet.cell(line, 7).value)       # energia Consumida no Contador em Ponta
        csp = (sheet.cell(line, 8).value)       # energia Consumida no Saldo em Ponta
        ccc = (sheet.cell(line, 9).value)       # energia Consumida no Contador em Cheias
        csc = (sheet.cell(line, 10).value)      # energia Consumida no Saldo em Cheias
        icv = (sheet.cell(line + 1, 5).value)   # energia Injetada no Contador em Vazio
        isv = (sheet.cell(line + 1, 6).value)   # energia Injetada no Saldo em Vazio
        icp = (sheet.cell(line + 1, 7).value)   # energia Injetada no Contador em Ponta
        isp = (sheet.cell(line + 1, 8).value)   # energia Injetada no Saldo em Ponta
        icc = (sheet.cell(line + 1, 9).value)   # energia Injetada no Contador em Cheias
        isc = (sheet.cell(line + 1, 10).value)  # energia Injetada no Saldo em Cheias

        line_of_data = [date, ccv, csv, ccp, csp, ccc, csc, icv, isv, icp, isp, icc, isc]
        data.append(line_of_data)
    
    return data


def new_database(directory):
    """Creates a new DataFrame given a folder of E-Redes .xlsx files.

    Args:
        directory (String): Path to the directory

    Returns:
        pd.Dataframe: A DataFrame with the values inside the folder
    """
    files = os.listdir(directory)
    data = []
    
    for file in files:
        data += open_reading(directory + "/" + file)
    
    df = pd.DataFrame(data, columns=["Date", "ccv", "csv", "ccp", "csp", "ccc", "csc",
                                     "icv", "isv", "icp", "isp", "icc", "isc"])
    
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True)
    for column in ["ccv", "csv", "ccp", "csp", "ccc", "csc", "icv", "isv", "icp", "isp", "icc", "isc"]:
        df[column] = pd.to_numeric(df[column], errors='coerce').astype("Int32")
    
    df.set_index("Date", inplace=True)
    df.sort_index(inplace=True, ascending=False)
    
    return df
    

if __name__ == "__main__":
    path = "/home/jvdcf/Projects/eredes-statistics/"
    df = new_database(path + "readings")
    print(df.info())
    df.to_csv(path + "database.csv", index=True)
